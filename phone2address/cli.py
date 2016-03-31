#!/usr/bin/env python
# -*- coding: utf-8 -*-


"""phone2address CLI

Usage:
    p2a <phone>
    p2a -f <file>
"""

from docopt import docopt
import openpyxl
import requests


def save_data(data, file):
    data.save(file)
    return True


def load_data(file):
    data = []
    wb = openpyxl.load_workbook(file).active
    i = 2
    while (wb.cell(row=i, column=1).value != None):
        data.append(str(int(wb.cell(row=i, column=1).value)))
        i += 1
    print(data)
    return data


def get_address(phone):
    url = 'http://apis.baidu.com/apistore/mobilenumber/mobilenumber?phone=%s' % (phone)
    header = {
        'apikey': 'beeabf9176db8aca915f178ced42c709'
    }
    r = requests.post(url, headers=header).json()
    if r['errNum'] != 0:
        return False
    else:
        return r['retData']


def process(file):
    data = load_data(file)
    wb = openpyxl.Workbook()
    ws = wb.active
    cur = 1
    for i in data:
        r = get_address(i)
        print(r)
        if r == False:
            ws.cell(row=cur, column=1).value = 'No a valid number'
        else:
            ws.cell(row=cur, column=1).value = r['phone']
            ws.cell(row=cur, column=2).value = r['province']
            ws.cell(row=cur, column=3).value = r['city']
            cur += 1
    save_data(wb, 'processed.xlsx')


def main():
    arguments = docopt(__doc__, version='0.0.1')
    if arguments['-f']:
        process(arguments['<file>'])
    else:
        r = get_address(arguments['phone'])
        if r == False:
            print('请输入有效的手机号码')
        else:
            print('您的手机归属地为%s%s' % (r['province'], r['city']))


if __name__ == '__main__':
    main()
